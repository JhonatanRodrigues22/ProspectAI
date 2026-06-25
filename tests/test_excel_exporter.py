import io

from openpyxl import load_workbook

from backend.app.domain import Lead, SearchResult
from backend.app.exporters import (
    export_search_result_csv,
    export_search_result_excel,
)
from backend.app.exporters.excel_exporter import WORKSHEET_TITLE
from backend.app.exporters.search_result_rows import EXPORT_HEADERS


def search_result(leads: list[Lead]) -> SearchResult:
    return SearchResult(
        origin_cep="01310100",
        category="cafeteria",
        radius_km=5,
        total_results=len(leads),
        leads=leads,
    )


def load_sheet(content: bytes):
    workbook = load_workbook(io.BytesIO(content))
    return workbook, workbook[WORKSHEET_TITLE]


def test_export_excel_with_leads() -> None:
    lead = Lead(
        name="Café São José",
        phone="+55 11 3333-4444",
        whatsapp="+55 11 99999-8888",
        email="contato@cafe.com",
        website="https://cafe.com",
        address="Avenida Paulista, 1000",
        city="São Paulo",
        state="SP",
        cep="01310100",
        rating=4.8,
        reviews_count=120,
        distance_km=1.234,
        source="google_places",
        source_id="place-123",
    )

    workbook, sheet = load_sheet(
        export_search_result_excel(search_result([lead]))
    )

    assert [cell.value for cell in sheet[1]] == EXPORT_HEADERS
    assert [cell.value for cell in sheet[2]] == [
        "Café São José",
        "+55 11 3333-4444",
        "+55 11 99999-8888",
        "contato@cafe.com",
        "https://cafe.com",
        "Avenida Paulista, 1000",
        "São Paulo",
        "SP",
        "01310100",
        4.8,
        120,
        1.234,
        "google_places",
        "place-123",
    ]
    workbook.close()


def test_export_empty_result_contains_headers_only() -> None:
    workbook, sheet = load_sheet(
        export_search_result_excel(search_result([]))
    )

    assert sheet.max_row == 1
    assert [cell.value for cell in sheet[1]] == EXPORT_HEADERS
    workbook.close()


def test_export_excel_writes_missing_fields_as_empty_cells() -> None:
    workbook, sheet = load_sheet(
        export_search_result_excel(
            search_result([Lead(name="Empresa")])
        )
    )

    assert [cell.value for cell in sheet[2]] == [
        "Empresa",
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        0,
        None,
        None,
        None,
    ]
    workbook.close()


def test_excel_file_can_be_opened_and_has_basic_formatting() -> None:
    content = export_search_result_excel(
        search_result([Lead(name="Empresa")])
    )

    assert content.startswith(b"PK")
    workbook, sheet = load_sheet(content)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:N2"
    assert sheet["A1"].font.bold is True
    workbook.close()


def test_csv_and_excel_exports_coexist() -> None:
    result = search_result([Lead(name="Empresa")])

    csv_content = export_search_result_csv(result)
    excel_content = export_search_result_excel(result)

    assert csv_content.startswith(b"\xef\xbb\xbf")
    assert excel_content.startswith(b"PK")
