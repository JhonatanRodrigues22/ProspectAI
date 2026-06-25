import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.app.domain import SearchResult
from backend.app.exporters.search_result_rows import (
    EXPORT_HEADERS,
    lead_export_row,
)

WORKSHEET_TITLE = "Leads"


def export_search_result_excel(result: SearchResult) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = WORKSHEET_TITLE
    worksheet.append(EXPORT_HEADERS)

    for lead in result.leads:
        worksheet.append(lead_export_row(lead))

    _format_worksheet(worksheet)

    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _format_worksheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    for column_cells in worksheet.columns:
        max_length = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        column_letter = get_column_letter(column_cells[0].column)
        worksheet.column_dimensions[column_letter].width = min(
            max(max_length + 2, 12),
            50,
        )
